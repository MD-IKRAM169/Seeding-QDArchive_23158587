from pathlib import Path
import sqlite3

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# PATHS
# ============================================================

PROJECT_ROOT = Path(__file__).resolve().parent.parent

CLASSIFICATION_DB = (
    PROJECT_ROOT / "23158587-sq26-classification.db"
)

EXPORT_DIR = PROJECT_ROOT / "exports"

OUTPUT_FILE = (
    EXPORT_DIR / "23158587-sq26-classification-results.xlsx"
)


# ============================================================
# LOAD RESULTS
# ============================================================

def load_results(
    conn: sqlite3.Connection,
) -> list[tuple]:
    """
    Load the required project-level classification results.

    Only QDA_PROJECT and QD_PROJECT records are included because
    these are the projects classified with ISIC Rev. 5.
    """

    rows = conn.execute(
        """
        SELECT
            p.repository_id,
            p.type AS project_type,
            p.title AS project_title,
            pc.primary_class,
            pc.secondary_class,

            (
                SELECT COUNT(*)
                FROM files AS f
                WHERE f.project_id = p.id
            ) AS no_project_files

        FROM projects AS p

        JOIN project_classifications AS pc
            ON pc.project_id = p.id

        WHERE p.type IN (
            'QDA_PROJECT',
            'QD_PROJECT'
        )

        ORDER BY
            p.repository_id,
            CASE p.type
                WHEN 'QDA_PROJECT' THEN 1
                WHEN 'QD_PROJECT' THEN 2
                ELSE 3
            END,
            p.id
        """
    ).fetchall()

    return rows


# ============================================================
# VERIFICATION
# ============================================================

def verify_results(
    conn: sqlite3.Connection,
    rows: list[tuple],
) -> None:
    """
    Verify the exported row count and required values.
    """

    expected_count = conn.execute(
        """
        SELECT COUNT(*)
        FROM projects
        WHERE type IN (
            'QDA_PROJECT',
            'QD_PROJECT'
        )
        """
    ).fetchone()[0]

    if len(rows) != expected_count:
        raise RuntimeError(
            "XLSX export verification failed. "
            f"Expected {expected_count} rows, "
            f"but found {len(rows)}."
        )

    invalid_rows = []

    for index, row in enumerate(
        rows,
        start=2,
    ):
        (
            repository_id,
            project_type,
            project_title,
            primary_class,
            secondary_class,
            no_project_files,
        ) = row

        if repository_id is None:
            invalid_rows.append(
                f"Row {index}: missing repository_id"
            )

        if project_type not in {
            "QDA_PROJECT",
            "QD_PROJECT",
        }:
            invalid_rows.append(
                f"Row {index}: invalid project_type"
            )

        if not project_title:
            invalid_rows.append(
                f"Row {index}: missing project_title"
            )

        if not primary_class:
            invalid_rows.append(
                f"Row {index}: missing primary_class"
            )

        if no_project_files is None:
            invalid_rows.append(
                f"Row {index}: missing file count"
            )

    if invalid_rows:
        raise RuntimeError(
            "\n".join(invalid_rows)
        )

    print(
        "Result verification successful."
    )

    print(
        f"Rows prepared for XLSX: {len(rows)}"
    )


# ============================================================
# CREATE XLSX
# ============================================================

def create_workbook(
    rows: list[tuple],
) -> Workbook:
    """
    Create and format the required workbook.
    """

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Classification Results"

    headers = [
        "repository_id",
        "project_type",
        "project_title",
        "primary_class",
        "secondary_class",
        "no_project_files",
    ]

    worksheet.append(headers)

    for row in rows:
        worksheet.append(row)

    # --------------------------------------------------------
    # HEADER STYLE
    # --------------------------------------------------------

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        bold=True,
        color="FFFFFF",
    )

    for cell in worksheet[1]:

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    # --------------------------------------------------------
    # GENERAL FORMATTING
    # --------------------------------------------------------

    worksheet.freeze_panes = "A2"

    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )

    for row in worksheet.iter_rows(
        min_row=2,
    ):

        for cell in row:

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    # Center selected columns.
    for column_letter in (
        "A",
        "B",
        "D",
        "E",
        "F",
    ):

        for cell in worksheet[
            column_letter
        ][1:]:

            cell.alignment = Alignment(
                horizontal="center",
                vertical="top",
                wrap_text=True,
            )

    # --------------------------------------------------------
    # COLUMN WIDTHS
    # --------------------------------------------------------

    column_widths = {
        "A": 15,
        "B": 18,
        "C": 70,
        "D": 16,
        "E": 18,
        "F": 18,
    }

    for (
        column_letter,
        width,
    ) in column_widths.items():

        worksheet.column_dimensions[
            column_letter
        ].width = width

    # Slightly taller header.
    worksheet.row_dimensions[1].height = 24

    return workbook


# ============================================================
# VERIFY WRITTEN XLSX
# ============================================================

def verify_workbook(
    workbook: Workbook,
    expected_rows: int,
) -> None:
    """
    Verify workbook structure before saving.
    """

    worksheet = workbook[
        "Classification Results"
    ]

    expected_headers = [
        "repository_id",
        "project_type",
        "project_title",
        "primary_class",
        "secondary_class",
        "no_project_files",
    ]

    actual_headers = [
        worksheet.cell(
            row=1,
            column=column_index,
        ).value

        for column_index in range(
            1,
            7,
        )
    ]

    if actual_headers != expected_headers:

        raise RuntimeError(
            "XLSX header verification failed."
        )

    actual_data_rows = (
        worksheet.max_row - 1
    )

    if actual_data_rows != expected_rows:

        raise RuntimeError(
            "XLSX row verification failed. "
            f"Expected {expected_rows}, "
            f"found {actual_data_rows}."
        )

    print(
        "Workbook verification successful."
    )


# ============================================================
# SHOW SUMMARY
# ============================================================

def show_summary(
    rows: list[tuple],
) -> None:
    """
    Show exported counts by repository and project type.
    """

    print("\n" + "=" * 72)

    print(
        "XLSX EXPORT SUMMARY"
    )

    print("=" * 72)

    counts = {}

    for row in rows:

        repository_id = row[0]

        project_type = row[1]

        key = (
            repository_id,
            project_type,
        )

        counts[key] = (
            counts.get(key, 0) + 1
        )

    for (
        repository_id,
        project_type,
    ), count in sorted(
        counts.items()
    ):

        print(
            f"Repository {repository_id}"
            f" | {project_type:<12}"
            f" | Rows: {count}"
        )

    print(
        f"\nTotal exported rows: "
        f"{len(rows)}"
    )


# ============================================================
# MAIN
# ============================================================

def main() -> None:
    """
    Main entry point.
    """

    print("=" * 72)

    print(
        "SEEDING QDARCHIVE - "
        "EXPORT CLASSIFICATION XLSX"
    )

    print("=" * 72)

    if not CLASSIFICATION_DB.exists():

        print("\nERROR:")

        print(
            "Classification database not found:"
        )

        print(CLASSIFICATION_DB)

        raise SystemExit(1)

    print(
        "\nClassification database:"
    )

    print(CLASSIFICATION_DB)

    try:

        with sqlite3.connect(
            CLASSIFICATION_DB
        ) as conn:

            rows = load_results(conn)

            verify_results(
                conn,
                rows,
            )

        workbook = create_workbook(
            rows
        )

        verify_workbook(
            workbook,
            len(rows),
        )

        EXPORT_DIR.mkdir(
            parents=True,
            exist_ok=True,
        )

        workbook.save(
            OUTPUT_FILE
        )

        show_summary(
            rows
        )

    except (
        sqlite3.Error,
        RuntimeError,
        OSError,
    ) as error:

        print("\nERROR:")

        print(error)

        raise SystemExit(1) from error

    print("\n" + "=" * 72)

    print("SUCCESS")

    print(
        "Classification XLSX exported successfully."
    )

    print(
        f"Output file:\n{OUTPUT_FILE}"
    )

    print("=" * 72)


if __name__ == "__main__":
    main()