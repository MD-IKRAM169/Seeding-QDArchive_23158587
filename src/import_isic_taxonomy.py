from pathlib import Path
import sqlite3

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parent.parent

CLASSIFICATION_DB = (
    PROJECT_ROOT / "23158587-sq26-classification.db"
)

ISIC_FILE = (
    PROJECT_ROOT / "ISIC5_Exp_Notes_11Mar2024.xlsx"
)

SHEET_NAME = "Divisions"


def create_isic_table(
    conn: sqlite3.Connection,
) -> None:
    """
    Create the ISIC division table if it does not exist.
    """

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS isic_divisions (
            code TEXT PRIMARY KEY,
            division_number INTEGER NOT NULL,
            title TEXT NOT NULL
        )
        """
    )


def read_isic_divisions() -> list[tuple[str, int, str]]:
    """
    Read ISIC division rows from the Excel workbook.
    """

    if not ISIC_FILE.exists():
        raise FileNotFoundError(
            f"ISIC Excel file not found:\n{ISIC_FILE}"
        )

    workbook = load_workbook(
        ISIC_FILE,
        data_only=True,
        read_only=True,
    )

    if SHEET_NAME not in workbook.sheetnames:
        raise RuntimeError(
            f"Worksheet '{SHEET_NAME}' not found."
        )

    worksheet = workbook[SHEET_NAME]

    divisions = []

    for row in worksheet.iter_rows(
        min_row=1,
        values_only=True,
    ):
        code = row[0]
        division_number = row[1]
        title = row[2]

        if not code or not title:
            continue

        normalized_code = str(code).strip()
        normalized_title = str(title).strip()

        try:
            normalized_number = int(
                float(division_number)
            )
        except (TypeError, ValueError):
            continue

        divisions.append(
            (
                normalized_code,
                normalized_number,
                normalized_title,
            )
        )

    return divisions


def import_divisions(
    conn: sqlite3.Connection,
    divisions: list[tuple[str, int, str]],
) -> None:
    """
    Insert or update all ISIC divisions.
    """

    conn.executemany(
        """
        INSERT INTO isic_divisions (
            code,
            division_number,
            title
        )
        VALUES (?, ?, ?)
        ON CONFLICT(code) DO UPDATE SET
            division_number = excluded.division_number,
            title = excluded.title
        """,
        divisions,
    )


def verify_import(
    conn: sqlite3.Connection,
) -> None:
    """
    Verify imported ISIC data.
    """

    count = conn.execute(
        """
        SELECT COUNT(*)
        FROM isic_divisions
        """
    ).fetchone()[0]

    print(f"\nImported ISIC divisions: {count}")

    if count != 87:
        raise RuntimeError(
            f"Expected 87 ISIC divisions, but found {count}."
        )

    print(
        "Verification successful: "
        "all 87 ISIC Rev. 5 divisions are present."
    )


def show_sample(
    conn: sqlite3.Connection,
) -> None:
    """
    Show sample ISIC records.
    """

    print("\nSample ISIC divisions:")

    rows = conn.execute(
        """
        SELECT
            code,
            title
        FROM isic_divisions
        ORDER BY code
        LIMIT 10
        """
    ).fetchall()

    for code, title in rows:
        print(f"  {code} - {title}")


def main() -> None:
    """
    Main entry point.
    """

    print("=" * 72)
    print(
        "SEEDING QDARCHIVE - "
        "IMPORT ISIC REV. 5 TAXONOMY"
    )
    print("=" * 72)

    if not CLASSIFICATION_DB.exists():
        print("\nERROR:")
        print("Classification database not found:")
        print(CLASSIFICATION_DB)
        raise SystemExit(1)

    print("\nClassification database:")
    print(CLASSIFICATION_DB)

    print("\nISIC source file:")
    print(ISIC_FILE)

    try:
        divisions = read_isic_divisions()

        print(
            f"\nDivision records read from Excel: "
            f"{len(divisions)}"
        )

        with sqlite3.connect(
            CLASSIFICATION_DB
        ) as conn:

            create_isic_table(conn)

            import_divisions(
                conn,
                divisions,
            )

            conn.commit()

            verify_import(conn)

            show_sample(conn)

    except (
        FileNotFoundError,
        RuntimeError,
        sqlite3.Error,
        OSError,
    ) as error:

        print("\nERROR:")
        print(error)

        raise SystemExit(1) from error

    print("\n" + "=" * 72)
    print("SUCCESS")
    print(
        "ISIC Rev. 5 division taxonomy "
        "imported successfully."
    )
    print("=" * 72)


if __name__ == "__main__":
    main()