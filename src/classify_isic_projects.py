from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
import html
import re
import sqlite3

from openpyxl import load_workbook
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity


# ============================================================
# PATHS
# ============================================================

PROJECT_ROOT = Path(__file__).resolve().parent.parent

CLASSIFICATION_DB = (
    PROJECT_ROOT / "23158587-sq26-classification.db"
)

ISIC_FILE = (
    PROJECT_ROOT / "ISIC5_Exp_Notes_11Mar2024.xlsx"
)

ISIC_SHEET = "ISIC5"


# ============================================================
# FIELD WEIGHTS
# ============================================================

# These sum to 1.0.
#
# Title, description, and keywords receive the highest weight
# because they usually describe the actual subject/domain of
# the project more directly than generic filenames.

TITLE_WEIGHT = 0.30
DESCRIPTION_WEIGHT = 0.30
KEYWORDS_WEIGHT = 0.20
FILE_CONTENT_WEIGHT = 0.15
FILE_NAMES_WEIGHT = 0.05


# ============================================================
# WORD/CHARACTER MODEL WEIGHTS
# ============================================================

WORD_MODEL_WEIGHT = 0.80
CHAR_MODEL_WEIGHT = 0.20


# ============================================================
# SECONDARY CLASS SETTINGS
# ============================================================

MIN_SECONDARY_SCORE = 0.035

# Secondary score must be at least 82% of primary score.
SECONDARY_RATIO = 0.82


# ============================================================
# CONFIDENCE SETTINGS
# ============================================================

LOW_ABSOLUTE_SCORE = 0.025
HIGH_ABSOLUTE_SCORE = 0.12

LOW_RELATIVE_MARGIN = 0.08
HIGH_RELATIVE_MARGIN = 0.25


# ============================================================
# GENERIC ACADEMIC WORDS
# ============================================================

# These words occur very frequently in research datasets and
# can create false matches with N72 simply because the object
# being classified is a research project.
#
# They are removed from project evidence before classification.

GENERIC_ACADEMIC_WORDS = {
    "research",
    "researcher",
    "researchers",
    "study",
    "studies",
    "studied",
    "data",
    "dataset",
    "datasets",
    "analysis",
    "analyses",
    "analyze",
    "analysed",
    "analyzed",
    "qualitative",
    "quantitative",
    "project",
    "projects",
    "interview",
    "interviews",
    "participant",
    "participants",
    "method",
    "methods",
    "methodology",
    "results",
    "result",
    "sample",
    "samples",
    "survey",
    "surveys",
    "questionnaire",
    "questionnaires",
    "paper",
    "article",
    "document",
    "documents",
    "file",
    "files",
    "repository",
    "metadata",
    "collection",
    "collected",
    "collecting",
}


# ============================================================
# TEXT CLEANING
# ============================================================

def clean_text(value: object | None) -> str:
    """
    Normalize text and whitespace.
    """

    if value is None:
        return ""

    text = html.unescape(str(value))

    text = text.replace("\x00", " ")

    text = re.sub(
        r"\s+",
        " ",
        text,
    )

    return text.strip()


def remove_generic_academic_words(
    text: str,
) -> str:
    """
    Remove generic academic terms that should not determine
    the economic/activity domain of a project.
    """

    if not text:
        return ""

    words = re.findall(
        r"\b[\w'-]+\b",
        text.lower(),
        flags=re.UNICODE,
    )

    filtered_words = [
        word
        for word in words
        if word not in GENERIC_ACADEMIC_WORDS
    ]

    return " ".join(filtered_words)


# ============================================================
# ISIC TAXONOMY
# ============================================================

def is_division_code(
    value: object | None,
) -> bool:
    """
    Identify division codes such as A01, B05, Q85, R86.
    """

    if value is None:
        return False

    code = str(value).strip().upper()

    return bool(
        re.fullmatch(
            r"[A-Z]\d{2}",
            code,
        )
    )


def read_isic_division_profiles() -> list[dict[str, str]]:
    """
    Read all 87 ISIC Rev. 5 division profiles.

    Positive evidence includes:
    - division title
    - introductory text
    - includes
    - includes also

    Exclusion text is deliberately not added as positive
    evidence because it describes activities belonging
    elsewhere.
    """

    if not ISIC_FILE.exists():
        raise FileNotFoundError(
            f"ISIC workbook not found:\n{ISIC_FILE}"
        )

    workbook = load_workbook(
        ISIC_FILE,
        read_only=True,
        data_only=True,
    )

    if ISIC_SHEET not in workbook.sheetnames:
        raise RuntimeError(
            f"Worksheet '{ISIC_SHEET}' not found."
        )

    worksheet = workbook[ISIC_SHEET]

    profiles = []

    for row in worksheet.iter_rows(
        min_row=2,
        values_only=True,
    ):
        code = row[0]
        title = row[2]
        introductory_text = row[3]
        includes = row[4]
        includes_also = row[5]

        if not is_division_code(code):
            continue

        normalized_code = (
            str(code)
            .strip()
            .upper()
        )

        normalized_title = clean_text(title)

        profile_parts = []

        if normalized_title:
            # Repeat title because division title is highly
            # important semantic evidence.
            profile_parts.extend(
                [
                    normalized_title,
                    normalized_title,
                    normalized_title,
                ]
            )

        for value in (
            introductory_text,
            includes,
            includes_also,
        ):
            cleaned = clean_text(value)

            if cleaned:
                profile_parts.append(cleaned)

        profile_text = " ".join(
            profile_parts
        )

        profiles.append(
            {
                "code": normalized_code,
                "title": normalized_title,
                "profile_text": profile_text,
            }
        )

    profiles.sort(
        key=lambda item: item["code"]
    )

    if len(profiles) != 87:
        raise RuntimeError(
            "Expected 87 ISIC Rev. 5 divisions, "
            f"but found {len(profiles)}."
        )

    return profiles


# ============================================================
# DATABASE TABLE
# ============================================================

def create_classification_table(
    conn: sqlite3.Connection,
) -> None:
    """
    Create detailed classification results table.
    """

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS project_classifications (
            project_id INTEGER PRIMARY KEY,
            primary_class TEXT NOT NULL,
            primary_title TEXT NOT NULL,
            primary_score REAL NOT NULL,
            secondary_class TEXT,
            secondary_title TEXT,
            secondary_score REAL,
            score_margin REAL NOT NULL,
            confidence TEXT NOT NULL,
            method TEXT NOT NULL,
            input_characters INTEGER NOT NULL,
            classified_at TEXT NOT NULL,
            FOREIGN KEY (project_id) REFERENCES projects(id)
        )
        """
    )


# ============================================================
# PROJECT INPUT PARSING
# ============================================================

def extract_labeled_value(
    text: str | None,
    label: str,
) -> str:
    """
    Extract one labeled field such as:

    TITLE: ...
    DESCRIPTION: ...
    LANGUAGE: ...
    """

    if not text:
        return ""

    pattern = (
        rf"(?:^|\n){re.escape(label)}:\s*"
        rf"(.*?)(?=\n[A-Z_ ]+:\s*|\Z)"
    )

    match = re.search(
        pattern,
        text,
        flags=re.DOTALL,
    )

    if not match:
        return ""

    return clean_text(
        match.group(1)
    )


def extract_keywords(
    metadata_text: str | None,
) -> str:
    """
    Extract the KEYWORDS field from prepared metadata.
    """

    return extract_labeled_value(
        metadata_text,
        "KEYWORDS",
    )


def load_projects(
    conn: sqlite3.Connection,
) -> list[dict[str, object]]:
    """
    Load prepared classification inputs separately by field.
    """

    rows = conn.execute(
        """
        SELECT
            ci.project_id,
            ci.repository_id,
            ci.project_type,
            p.title,
            p.description,
            ci.metadata_text,
            ci.file_names_text,
            ci.file_content_text,
            ci.combined_text
        FROM classification_inputs AS ci
        JOIN projects AS p
            ON p.id = ci.project_id
        WHERE ci.project_type IN (
            'QDA_PROJECT',
            'QD_PROJECT'
        )
        ORDER BY ci.project_id
        """
    ).fetchall()

    projects = []

    for (
        project_id,
        repository_id,
        project_type,
        title,
        description,
        metadata_text,
        file_names_text,
        file_content_text,
        combined_text,
    ) in rows:

        cleaned_title = remove_generic_academic_words(
            clean_text(title)
        )

        cleaned_description = remove_generic_academic_words(
            clean_text(description)
        )

        keywords = remove_generic_academic_words(
            extract_keywords(metadata_text)
        )

        filenames = remove_generic_academic_words(
            clean_text(file_names_text)
        )

        file_content = remove_generic_academic_words(
            clean_text(file_content_text)
        )

        combined = clean_text(combined_text)

        if not combined:
            raise RuntimeError(
                "Empty classification input for "
                f"project {project_id}."
            )

        projects.append(
            {
                "project_id": project_id,
                "repository_id": repository_id,
                "project_type": project_type,
                "title": cleaned_title,
                "description": cleaned_description,
                "keywords": keywords,
                "file_names": filenames,
                "file_content": file_content,
                "combined_text": combined,
            }
        )

    return projects


# ============================================================
# SIMILARITY MODEL
# ============================================================

def calculate_field_similarity(
    project_texts: list[str],
    class_texts: list[str],
):
    """
    Calculate hybrid word + character TF-IDF similarity
    for one specific evidence field.
    """

    complete_corpus = (
        class_texts
        + project_texts
    )

    # Word model
    word_vectorizer = TfidfVectorizer(
        lowercase=True,
        strip_accents="unicode",
        stop_words="english",
        ngram_range=(1, 2),
        min_df=1,
        sublinear_tf=True,
        max_features=60_000,
    )

    word_matrix = (
        word_vectorizer
        .fit_transform(complete_corpus)
    )

    class_word_matrix = word_matrix[
        :len(class_texts)
    ]

    project_word_matrix = word_matrix[
        len(class_texts):
    ]

    word_similarity = cosine_similarity(
        project_word_matrix,
        class_word_matrix,
    )

    # Character model
    char_vectorizer = TfidfVectorizer(
        analyzer="char_wb",
        lowercase=True,
        strip_accents="unicode",
        ngram_range=(3, 5),
        min_df=1,
        sublinear_tf=True,
        max_features=60_000,
    )

    char_matrix = (
        char_vectorizer
        .fit_transform(complete_corpus)
    )

    class_char_matrix = char_matrix[
        :len(class_texts)
    ]

    project_char_matrix = char_matrix[
        len(class_texts):
    ]

    char_similarity = cosine_similarity(
        project_char_matrix,
        class_char_matrix,
    )

    return (
        WORD_MODEL_WEIGHT * word_similarity
        + CHAR_MODEL_WEIGHT * char_similarity
    )


def build_similarity_matrix(
    class_profiles: list[dict[str, str]],
    projects: list[dict[str, object]],
):
    """
    Score each project using separate evidence fields.

    Final score:
    30% title
    30% description
    20% keywords
    15% extracted file content
     5% useful filenames

    Missing fields automatically contribute zero.
    """

    class_texts = [
        item["profile_text"]
        for item in class_profiles
    ]

    fields = [
        (
            "title",
            TITLE_WEIGHT,
        ),
        (
            "description",
            DESCRIPTION_WEIGHT,
        ),
        (
            "keywords",
            KEYWORDS_WEIGHT,
        ),
        (
            "file_content",
            FILE_CONTENT_WEIGHT,
        ),
        (
            "file_names",
            FILE_NAMES_WEIGHT,
        ),
    ]

    final_similarity = None

    print(
        "\nBuilding weighted field-level "
        "similarity models..."
    )

    for field_name, field_weight in fields:

        project_texts = [
            str(
                project.get(
                    field_name,
                    "",
                )
            )
            for project in projects
        ]

        nonempty_count = sum(
            1
            for text in project_texts
            if text.strip()
        )

        print(
            f"  {field_name:<14}"
            f" | weight={field_weight:.2f}"
            f" | non-empty projects="
            f"{nonempty_count}"
        )

        field_similarity = (
            calculate_field_similarity(
                project_texts,
                class_texts,
            )
        )

        weighted_similarity = (
            field_weight
            * field_similarity
        )

        if final_similarity is None:
            final_similarity = (
                weighted_similarity
            )
        else:
            final_similarity += (
                weighted_similarity
            )

    return final_similarity


# ============================================================
# RESULT SELECTION
# ============================================================

def determine_confidence(
    primary_score: float,
    secondary_score: float,
) -> str:
    """
    Determine confidence using both absolute score and
    relative separation between first and second place.
    """

    if primary_score <= 0:
        return "LOW"

    margin = (
        primary_score - secondary_score
    )

    relative_margin = (
        margin / primary_score
    )

    if (
        primary_score >= HIGH_ABSOLUTE_SCORE
        and relative_margin >= HIGH_RELATIVE_MARGIN
    ):
        return "HIGH"

    if (
        primary_score < LOW_ABSOLUTE_SCORE
        or relative_margin < LOW_RELATIVE_MARGIN
    ):
        return "LOW"

    return "MEDIUM"


def choose_classes(
    scores,
    class_profiles: list[dict[str, str]],
) -> dict[str, object]:
    """
    Select primary and optional secondary ISIC classes.
    """

    ranked_indexes = sorted(
        range(len(scores)),
        key=lambda index: float(
            scores[index]
        ),
        reverse=True,
    )

    primary_index = ranked_indexes[0]
    second_index = ranked_indexes[1]

    primary_score = float(
        scores[primary_index]
    )

    second_score = float(
        scores[second_index]
    )

    primary_profile = class_profiles[
        primary_index
    ]

    second_profile = class_profiles[
        second_index
    ]

    score_margin = (
        primary_score - second_score
    )

    use_secondary = (
        second_score >= MIN_SECONDARY_SCORE
        and primary_score > 0
        and (
            second_score / primary_score
        ) >= SECONDARY_RATIO
    )

    confidence = determine_confidence(
        primary_score,
        second_score,
    )

    result = {
        "primary_class": (
            primary_profile["code"]
        ),
        "primary_title": (
            primary_profile["title"]
        ),
        "primary_score": primary_score,
        "secondary_class": None,
        "secondary_title": None,
        "secondary_score": None,
        "score_margin": score_margin,
        "confidence": confidence,
    }

    if use_secondary:
        result["secondary_class"] = (
            second_profile["code"]
        )

        result["secondary_title"] = (
            second_profile["title"]
        )

        result["secondary_score"] = (
            second_score
        )

    return result


# ============================================================
# STORE RESULTS
# ============================================================

def store_result(
    conn: sqlite3.Connection,
    project: dict[str, object],
    result: dict[str, object],
) -> None:
    """
    Store result and update projects.class.
    """

    timestamp = datetime.now(
        timezone.utc
    ).isoformat()

    method = (
        "weighted_field_hybrid_tfidf_v2_"
        "isic_rev5_division"
    )

    conn.execute(
        """
        INSERT INTO project_classifications (
            project_id,
            primary_class,
            primary_title,
            primary_score,
            secondary_class,
            secondary_title,
            secondary_score,
            score_margin,
            confidence,
            method,
            input_characters,
            classified_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)

        ON CONFLICT(project_id) DO UPDATE SET
            primary_class = excluded.primary_class,
            primary_title = excluded.primary_title,
            primary_score = excluded.primary_score,
            secondary_class = excluded.secondary_class,
            secondary_title = excluded.secondary_title,
            secondary_score = excluded.secondary_score,
            score_margin = excluded.score_margin,
            confidence = excluded.confidence,
            method = excluded.method,
            input_characters = excluded.input_characters,
            classified_at = excluded.classified_at
        """,
        (
            project["project_id"],
            result["primary_class"],
            result["primary_title"],
            result["primary_score"],
            result["secondary_class"],
            result["secondary_title"],
            result["secondary_score"],
            result["score_margin"],
            result["confidence"],
            method,
            len(
                str(
                    project[
                        "combined_text"
                    ]
                )
            ),
            timestamp,
        ),
    )

    conn.execute(
        """
        UPDATE projects
        SET class = ?
        WHERE id = ?
        """,
        (
            result["primary_class"],
            project["project_id"],
        ),
    )


# ============================================================
# VERIFICATION
# ============================================================

def verify_results(
    conn: sqlite3.Connection,
    expected_count: int,
) -> None:
    """
    Verify all relevant projects were classified.
    """

    actual_count = conn.execute(
        """
        SELECT COUNT(*)
        FROM project_classifications
        """
    ).fetchone()[0]

    if actual_count != expected_count:
        raise RuntimeError(
            "Expected "
            f"{expected_count} classifications, "
            f"but found {actual_count}."
        )

    missing_count = conn.execute(
        """
        SELECT COUNT(*)
        FROM projects
        WHERE type IN (
            'QDA_PROJECT',
            'QD_PROJECT'
        )
        AND (
            class IS NULL
            OR TRIM(class) = ''
        )
        """
    ).fetchone()[0]

    if missing_count:
        raise RuntimeError(
            f"{missing_count} relevant projects "
            "have no class."
        )

    print("\nVerification successful.")

    print(
        f"Classified projects: "
        f"{actual_count}"
    )

    print(
        "Relevant projects without class: "
        f"{missing_count}"
    )


# ============================================================
# SUMMARY
# ============================================================

def show_summary(
    conn: sqlite3.Connection,
) -> None:
    """
    Show classification statistics.
    """

    print("\n" + "=" * 78)
    print("CLASSIFICATION SUMMARY")
    print("=" * 78)

    rows = conn.execute(
        """
        SELECT
            p.repository_id,
            p.type,
            COUNT(*)
        FROM project_classifications AS pc
        JOIN projects AS p
            ON p.id = pc.project_id
        GROUP BY
            p.repository_id,
            p.type
        ORDER BY
            p.repository_id,
            p.type
        """
    ).fetchall()

    for (
        repository_id,
        project_type,
        count,
    ) in rows:

        print(
            f"Repository {repository_id}"
            f" | {project_type:<12}"
            f" | Projects: {count}"
        )

    print("\nConfidence levels:")

    confidence_rows = conn.execute(
        """
        SELECT
            confidence,
            COUNT(*)
        FROM project_classifications
        GROUP BY confidence
        ORDER BY
            CASE confidence
                WHEN 'HIGH' THEN 1
                WHEN 'MEDIUM' THEN 2
                WHEN 'LOW' THEN 3
                ELSE 4
            END
        """
    ).fetchall()

    for confidence, count in confidence_rows:
        print(
            f"  {confidence:<6}: {count}"
        )

    print("\nMost common primary classes:")

    class_rows = conn.execute(
        """
        SELECT
            primary_class,
            primary_title,
            COUNT(*) AS class_count
        FROM project_classifications
        GROUP BY
            primary_class,
            primary_title
        ORDER BY
            class_count DESC,
            primary_class
        LIMIT 20
        """
    ).fetchall()

    for code, title, count in class_rows:
        print(
            f"  {code:<4}"
            f" | {count:>2}"
            f" | {title}"
        )


# ============================================================
# MAIN
# ============================================================

def main() -> None:
    """
    Main entry point.
    """

    print("=" * 78)
    print(
        "SEEDING QDARCHIVE - "
        "IMPROVED ISIC REV. 5 PROJECT CLASSIFIER"
    )
    print("=" * 78)

    if not CLASSIFICATION_DB.exists():
        print("\nERROR:")
        print(
            "Classification database not found:"
        )
        print(CLASSIFICATION_DB)
        raise SystemExit(1)

    if not ISIC_FILE.exists():
        print("\nERROR:")
        print(
            "ISIC workbook not found:"
        )
        print(ISIC_FILE)
        raise SystemExit(1)

    print("\nClassification database:")
    print(CLASSIFICATION_DB)

    print("\nISIC workbook:")
    print(ISIC_FILE)

    try:
        print(
            "\nReading ISIC Rev. 5 "
            "division profiles..."
        )

        class_profiles = (
            read_isic_division_profiles()
        )

        print(
            "ISIC division profiles loaded: "
            f"{len(class_profiles)}"
        )

        with sqlite3.connect(
            CLASSIFICATION_DB
        ) as conn:

            create_classification_table(
                conn
            )

            projects = load_projects(conn)

            print(
                "\nPrepared projects to classify: "
                f"{len(projects)}"
            )

            if not projects:
                raise RuntimeError(
                    "No prepared classification "
                    "inputs found."
                )

            similarities = (
                build_similarity_matrix(
                    class_profiles,
                    projects,
                )
            )

            # Clear previous results.
            conn.execute(
                """
                DELETE FROM project_classifications
                """
            )

            print(
                "\nClassifying projects..."
            )

            for index, project in enumerate(
                projects
            ):
                result = choose_classes(
                    similarities[index],
                    class_profiles,
                )

                store_result(
                    conn,
                    project,
                    result,
                )

                secondary_display = (
                    result[
                        "secondary_class"
                    ]
                    or "-"
                )

                print(
                    f"Project "
                    f"{project['project_id']:>3}"
                    f" | Repo "
                    f"{project['repository_id']}"
                    f" | "
                    f"{project['project_type']:<11}"
                    f" | Primary: "
                    f"{result['primary_class']}"
                    f" ({result['primary_score']:.4f})"
                    f" | Secondary: "
                    f"{secondary_display}"
                    f" | Confidence: "
                    f"{result['confidence']}"
                )

            conn.commit()

            verify_results(
                conn,
                len(projects),
            )

            show_summary(conn)

    except (
        FileNotFoundError,
        RuntimeError,
        sqlite3.Error,
        OSError,
        ValueError,
    ) as error:

        print("\nERROR:")
        print(error)

        raise SystemExit(1) from error

    print("\n" + "=" * 78)
    print("SUCCESS")
    print(
        "Improved ISIC Rev. 5 project "
        "classification completed."
    )
    print("=" * 78)


if __name__ == "__main__":
    main()